from openpyxl import load_workbook
from datetime import datetime
import json
# .xlsx dosyasını yükleme
workbook = load_workbook(filename=r"C:\Users\mertg\OneDrive\Masaüstü\Dosya\Programming\web\random-proje\lineup_maker\xlsx-to-json\data.xlsx")

# Aktif çalışma sayfasını seçme
sheet = workbook.active

def calculate_age(birth_date):
    today = datetime.today()
    age = today.year - birth_date.year
    
    # Eğer doğum gününü henüz kutlamadıysa yaşı bir yıl azalt
    if (today.month, today.day) < (birth_date.month, birth_date.day):
        age -= 1
    
    return age

# Belirli bir hücreyi okuma
cell_value = sheet['A1'].value
print(f'A1 hücresinin değeri: {cell_value}')
tarih = None

total_data={"players":[]}
x=0
# Tüm satırları ve sütunları okuma
for row in sheet.iter_rows(values_only=True):
    satir = list(row)
    tarih=satir[10]
    yas=None
    if x==0:
        x+=1
        continue
    try:
        yas=calculate_age(satir[10])
    except:
        print("tarih yok")
    if yas==None:
        data={
            "name":satir[2],
            "position":satir[4],
            "rating":satir[5],
            "age":str(int(satir[9]) + 2),
            "country":satir[23],
            "birth_date":satir[10].strftime('%m/%d/%Y') if satir[10]!=None else None,
            "country_flag":satir[-1],
            "photo":satir[-5]
        }
    else:
        data={
            "name":satir[2],
            "position":satir[4],
            "rating":satir[5],
            "age":yas,
            "country":satir[23],
            "birth_date":satir[10].strftime('%m/%d/%Y') if satir[10]!=None else None,
            "country_flag":satir[-1],
            "photo":satir[-5]
        }
    
    
    total_data["players"].append(data)


with open('./data.json', 'w', encoding='utf-8') as json_file:
    json.dump(total_data, json_file, ensure_ascii=False, indent=4)





